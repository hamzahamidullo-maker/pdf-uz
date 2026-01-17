import os
import json
import asyncio
import sqlite3
from datetime import datetime, date
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ContextTypes, filters, CallbackQueryHandler
)
from PIL import Image
from fpdf import FPDF
import docx
import openpyxl
from flask import Flask, request


# ===== DATABASE SETUP =====
def init_database():
    """Ma'lumotlar bazasini ishga tushirish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    # Foydalanuvchilar jadvali
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS users (
        user_id INTEGER PRIMARY KEY,
        username TEXT,
        first_name TEXT,
        last_name TEXT,
        created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        last_active TIMESTAMP,
        total_pdfs INTEGER DEFAULT 0,
        total_files INTEGER DEFAULT 0,
        is_admin BOOLEAN DEFAULT 0
    )
    ''')
    
    # Statistika jadvali
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS stats (
        date DATE PRIMARY KEY,
        new_users INTEGER DEFAULT 0,
        active_users INTEGER DEFAULT 0,
        total_pdfs INTEGER DEFAULT 0,
        total_files INTEGER DEFAULT 0
    )
    ''')
    
    # Adminlar ro'yxati
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS admins (
        user_id INTEGER PRIMARY KEY,
        username TEXT,
        added_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    conn.commit()
    conn.close()

# ===== CONFIG =====
with open("data/config.json", "r", encoding="utf-8") as f:
    config = json.load(f)

TOKEN = os.getenv("BOT_TOKEN")

CHANNEL_USERNAME = config.get("CHANNEL_USERNAME", "@test_channel")
ADMIN_IDS = config.get("ADMIN_IDS", [])  # Admin ID larini config dan olish

os.makedirs("temp", exist_ok=True)
os.makedirs("data", exist_ok=True)

# Ma'lumotlar bazasini ishga tushirish
init_database()

# ===== USER FILES =====
user_files = {}
user_pdf_counter = {}
user_tasks = {}
user_subscribed = {}
user_progress_msg_id = {}

# ===== DATABASE FUNCTIONS =====
def add_user(user_id, username, first_name, last_name):
    """Yangi foydalanuvchi qo'shish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    INSERT OR IGNORE INTO users (user_id, username, first_name, last_name, last_active)
    VALUES (?, ?, ?, ?, datetime('now'))
    ''', (user_id, username, first_name, last_name))
    
    conn.commit()
    conn.close()

def update_user_activity(user_id):
    """Foydalanuvchi faolligini yangilash"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    UPDATE users 
    SET last_active = datetime('now')
    WHERE user_id = ?
    ''', (user_id,))
    
    conn.commit()
    conn.close()

def increment_user_stats(user_id, pdfs=0, files=0):
    """Foydalanuvchi statistikasini oshirish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    UPDATE users 
    SET total_pdfs = total_pdfs + ?,
        total_files = total_files + ?
    WHERE user_id = ?
    ''', (pdfs, files, user_id))
    
    conn.commit()
    conn.close()

def update_daily_stats():
    """Kunlik statistikani yangilash"""
    today = date.today().isoformat()
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    # Bugungi sana uchun yozuv yo'q bo'lsa yaratish
    cursor.execute('''
    INSERT OR IGNORE INTO stats (date) VALUES (?)
    ''', (today,))
    
    # Faol foydalanuvchilar sonini hisoblash (oxirgi 24 soatda faol)
    cursor.execute('''
    SELECT COUNT(*) FROM users 
    WHERE last_active > datetime('now', '-1 day')
    ''')
    active_users = cursor.fetchone()[0]
    
    # Yangi foydalanuvchilar sonini hisoblash (bugun qo'shilgan)
    cursor.execute('''
    SELECT COUNT(*) FROM users 
    WHERE DATE(created_date) = DATE('now')
    ''')
    new_users = cursor.fetchone()[0]
    
    # Bugungi PDF va fayllar sonini hisoblash
    cursor.execute('''
    SELECT SUM(total_pdfs), SUM(total_files) 
    FROM users 
    WHERE DATE(last_active) = DATE('now')
    ''')
    daily_stats = cursor.fetchone()
    daily_pdfs = daily_stats[0] or 0
    daily_files = daily_stats[1] or 0
    
    # Statistikani yangilash
    cursor.execute('''
    UPDATE stats 
    SET new_users = ?, 
        active_users = ?,
        total_pdfs = ?,
        total_files = ?
    WHERE date = ?
    ''', (new_users, active_users, daily_pdfs, daily_files, today))
    
    conn.commit()
    conn.close()

def get_bot_stats():
    """Bot umumiy statistikasini olish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    # Umumiy foydalanuvchilar soni
    cursor.execute('SELECT COUNT(*) FROM users')
    total_users = cursor.fetchone()[0]
    
    # Bugungi faol foydalanuvchilar
    cursor.execute('''
    SELECT COUNT(*) FROM users 
    WHERE last_active > datetime('now', '-1 day')
    ''')
    today_active = cursor.fetchone()[0]
    
    # Bugungi yangi foydalanuvchilar
    cursor.execute('''
    SELECT COUNT(*) FROM users 
    WHERE DATE(created_date) = DATE('now')
    ''')
    today_new = cursor.fetchone()[0]
    
    # Umumiy PDF lar soni
    cursor.execute('SELECT SUM(total_pdfs) FROM users')
    total_pdfs = cursor.fetchone()[0] or 0
    
    # Umumiy fayllar soni
    cursor.execute('SELECT SUM(total_files) FROM users')
    total_files = cursor.fetchone()[0] or 0
    
    conn.close()
    
    return {
        'total_users': total_users,
        'today_active': today_active,
        'today_new': today_new,
        'total_pdfs': total_pdfs,
        'total_files': total_files
    }

def get_user_stats(user_id=None):
    """Foydalanuvchi statistikasini olish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    if user_id:
        cursor.execute('''
        SELECT user_id, username, first_name, last_name, 
               created_date, last_active, total_pdfs, total_files
        FROM users WHERE user_id = ?
        ''', (user_id,))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            return {
                'user_id': user[0],
                'username': user[1],
                'first_name': user[2],
                'last_name': user[3],
                'created_date': user[4],
                'last_active': user[5],
                'total_pdfs': user[6],
                'total_files': user[7]
            }
        return None
    
    else:
        # Eng faol 10 ta foydalanuvchi
        cursor.execute('''
        SELECT user_id, username, first_name, last_active, total_pdfs, total_files
        FROM users 
        ORDER BY last_active DESC 
        LIMIT 10
        ''')
        active_users = cursor.fetchall()
        
        # Eng ko'p PDF yaratgan 10 ta foydalanuvchi
        cursor.execute('''
        SELECT user_id, username, first_name, total_pdfs, total_files
        FROM users 
        ORDER BY total_pdfs DESC 
        LIMIT 10
        ''')
        top_pdf_users = cursor.fetchall()
        
        conn.close()
        
        return {
            'active_users': active_users,
            'top_pdf_users': top_pdf_users
        }

def get_daily_stats(days=7):
    """Oxirgi n kunlik statistikani olish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    SELECT date, new_users, active_users, total_pdfs, total_files
    FROM stats 
    WHERE date >= date('now', ?)
    ORDER BY date DESC
    ''', (f'-{days} days',))
    
    stats = cursor.fetchall()
    conn.close()
    
    return stats

def add_admin(user_id, username):
    """Admin qo'shish"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('''
    INSERT OR IGNORE INTO admins (user_id, username)
    VALUES (?, ?)
    ''', (user_id, username))
    
    # users jadvalida ham admin sifatida belgilash
    cursor.execute('''
    UPDATE users SET is_admin = 1 WHERE user_id = ?
    ''', (user_id,))
    
    conn.commit()
    conn.close()

def remove_admin(user_id):
    """Adminni olib tashlash"""
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('DELETE FROM admins WHERE user_id = ?', (user_id,))
    
    # users jadvalidan admin holatini olib tashlash
    cursor.execute('UPDATE users SET is_admin = 0 WHERE user_id = ?', (user_id,))
    
    conn.commit()
    conn.close()

def is_admin(user_id):
    """Foydalanuvchi admin ekanligini tekshirish"""
    # Birinchi config dan tekshirish
    if user_id in ADMIN_IDS:
        return True
    
    # Keyin ma'lumotlar bazasidan tekshirish
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    
    cursor.execute('SELECT 1 FROM admins WHERE user_id = ?', (user_id,))
    result = cursor.fetchone()
    
    conn.close()
    
    return result is not None

# ===== KANAL MA'LUMOTLARINI OLISH =====
async def get_channel_info(context: ContextTypes.DEFAULT_TYPE):
    try:
        channel_username = CHANNEL_USERNAME.lstrip('@')
        try:
            chat = await context.bot.get_chat(chat_id=f"@{channel_username}")
        except:
            chat = await context.bot.get_chat(chat_id=channel_username)
        
        channel_title = chat.title
        channel_link = f"https://t.me/{channel_username}"
        
        return {
            "title": channel_title,
            "username": CHANNEL_USERNAME,
            "link": channel_link,
            "chat_id": chat.id
        }
        
    except Exception as e:
        print(f"Kanal ma'lumotlarini olish xatosi: {e}")
        return {
            "title": CHANNEL_USERNAME,
            "username": CHANNEL_USERNAME,
            "link": f"https://t.me/{CHANNEL_USERNAME.lstrip('@')}",
            "chat_id": None
        }

# ===== KANALGA OBUNA BO'LGANLIGINI TEKSHIRISH =====
async def check_subscription(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
    try:
        channel_info = await get_channel_info(context)
        
        if not channel_info.get("chat_id"):
            return True
        
        chat_member = await context.bot.get_chat_member(
            chat_id=channel_info["chat_id"],
            user_id=user_id
        )
        
        status = chat_member.status
        return status in ['member', 'administrator', 'creator']
        
    except Exception as e:
        print(f"Obuna tekshirish xatosi: {e}")
        return True

# ===== /start =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    user_id = u.id
    
    # Foydalanuvchini ma'lumotlar bazasiga qo'shish
    add_user(user_id, u.username, u.first_name, u.last_name)
    update_user_activity(user_id)
    
    # Barcha eski ma'lumotlarni tozalash
    user_files[user_id] = []
    user_pdf_counter[user_id] = 1
    if user_id in user_tasks:
        try:
            user_tasks[user_id].cancel()
        except:
            pass
        user_tasks.pop(user_id, None)
    
    # Progress xabarni tozalash
    user_progress_msg_id.pop(user_id, None)
    
    # Kanal ma'lumotlarini olish
    channel_info = await get_channel_info(context)
    
    # Obunani tekshirish
    is_subscribed = await check_subscription(user_id, context)
    
    if not is_subscribed:
        keyboard = [
            [InlineKeyboardButton(f"📢 {channel_info['title']}", url=channel_info["link"])],
            [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_subscription")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f"👋 Assalomu alaykum {u.first_name}!\n\n"
            f"📄 **PDF Converter Bot** ga xush kelibsiz!\n\n"
            f"Botdan foydalanish uchun quyidagi kanalga obuna bo'lishingiz kerak:\n"
            f"📢 **{channel_info['title']}**\n"
            f"📍 {channel_info['username']}\n\n"
            f"Kanalga obuna bo'lgach, '✅ Obunani tekshirish' tugmasini bosing.",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return
    
    # Obuna bo'lgan foydalanuvchi uchun
    user_subscribed[user_id] = True
    
    welcome_text = f"""
✅ **Xush kelibsiz {u.first_name}!**

🤖 **PDF Converter Bot**

📤 **Fayllarni yuboring va PDF oling!**
    
⚡️ **Tezkor va oddiy:**
1. Rasm, Word yoki Excel fayllarini yuboring
2. 3 soniya kuting
3. PDF avtomatik yuboriladi

📊 **Har safar yangi PDF yaratiladi**
"""
    
    await update.message.reply_text(
        welcome_text,
        parse_mode='Markdown'
    )

# ===== ADMIN PANEL =====
async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin panelini ko'rsatish"""
    user_id = update.effective_user.id
    
    # Admin ekanligini tekshirish
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    # Statistikani yangilash
    update_daily_stats()
    
    # Bot statistikasini olish
    stats = get_bot_stats()
    
    # Kunlik statistikani olish
    daily_stats = get_daily_stats(7)
    
    # Admin panel keyboard
    keyboard = [
        [InlineKeyboardButton("📊 Umumiy statistika", callback_data="admin_stats")],
        [InlineKeyboardButton("👥 Foydalanuvchilar", callback_data="admin_users")],
        [InlineKeyboardButton("📈 Kunlik statistika", callback_data="admin_daily")],
        [InlineKeyboardButton("⚙️ Admin boshqaruvi", callback_data="admin_manage")],
        [InlineKeyboardButton("📤 Ma'lumotlarni yuklash", callback_data="admin_export")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    stats_text = f"""
🔧 **ADMIN PANELI**

👥 **Foydalanuvchilar:**
• Jami obunachilar: {stats['total_users']}
• Bugungi faol: {stats['today_active']}
• Bugungi yangi: {stats['today_new']}

📄 **PDF Statistika:**
• Jami PDF lar: {stats['total_pdfs']}
• Jami fayllar: {stats['total_files']}

📅 **Oxirgi 7 kun:**
"""
    
    # Oxirgi 7 kunlik statistikani qo'shish
    for day in daily_stats[:5]:  # Faqat oxirgi 5 kunni ko'rsatish
        day_date = datetime.strptime(day[0], '%Y-%m-%d').strftime('%d.%m')
        stats_text += f"• {day_date}: {day[2]} faol, {day[1]} yangi\n"
    
    await update.message.reply_text(
        stats_text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def admin_stats_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Umumiy statistika callback"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.edit_message_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    # Statistikani yangilash
    update_daily_stats()
    
    # Bot statistikasini olish
    stats = get_bot_stats()
    
    # Eng faol foydalanuvchilar
    user_stats = get_user_stats()
    
    stats_text = f"""
📊 **UMUMIY STATISTIKA**

👥 **Foydalanuvchilar:**
• Jami obunachilar: {stats['total_users']}
• Bugungi faol: {stats['today_active']}
• Bugungi yangi: {stats['today_new']}

📄 **PDF Statistika:**
• Jami PDF lar: {stats['total_pdfs']}
• Jami fayllar: {stats['total_files']}

🏆 **Eng faol foydalanuvchilar (oxirgi 24 soat):**
"""
    
    # Eng faol foydalanuvchilarni qo'shish
    for i, user in enumerate(user_stats['active_users'][:5], 1):
        username = user[1] or f"{user[2]}"
        last_active = datetime.strptime(user[3], '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
        stats_text += f"{i}. {username} - {last_active}\n"
    
    stats_text += "\n🏅 **Eng ko'p PDF yaratganlar:**\n"
    
    # Eng ko'p PDF yaratganlarni qo'shish
    for i, user in enumerate(user_stats['top_pdf_users'][:5], 1):
        username = user[1] or f"{user[2]}"
        stats_text += f"{i}. {username} - {user[3]} PDF\n"
    
    # Orqaga qaytish tugmasi
    keyboard = [[InlineKeyboardButton("🔙 Orqaga", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        stats_text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def admin_users_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchilar ro'yxati callback"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.edit_message_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    # Foydalanuvchi statistikasini olish
    user_stats = get_user_stats()
    
    users_text = "👥 **FOYDALANUVCHILAR RO'YXATI**\n\n"
    users_text += "🕐 **Eng faol foydalanuvchilar:**\n"
    
    # Eng faol foydalanuvchilar
    for i, user in enumerate(user_stats['active_users'][:10], 1):
        username = user[1] or f"{user[2]}"
        last_active = datetime.strptime(user[3], '%Y-%m-%d %H:%M:%S').strftime('%d.%m %H:%M')
        users_text += f"{i}. {username} - {last_active}\n"
    
    users_text += "\n📄 **Eng ko'p PDF yaratganlar:**\n"
    
    # Eng ko'p PDF yaratganlar
    for i, user in enumerate(user_stats['top_pdf_users'][:10], 1):
        username = user[1] or f"{user[2]}"
        users_text += f"{i}. {username} - {user[3]} PDF, {user[4]} fayl\n"
    
    # Orqaga qaytish tugmasi
    keyboard = [[InlineKeyboardButton("🔙 Orqaga", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        users_text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def admin_daily_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kunlik statistika callback"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.edit_message_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    # Kunlik statistikani olish
    daily_stats = get_daily_stats(14)  # Oxirgi 14 kun
    
    daily_text = "📈 **KUNLIK STATISTIKA (Oxirgi 14 kun)**\n\n"
    daily_text += "📅 Sana | Yangi | Faol | PDF | Fayllar\n"
    daily_text += "─" * 40 + "\n"
    
    for day in daily_stats:
        day_date = datetime.strptime(day[0], '%Y-%m-%d').strftime('%d.%m')
        daily_text += f"{day_date} | {day[1]} | {day[2]} | {day[3]} | {day[4]}\n"
    
    # Orqaga qaytish tugmasi
    keyboard = [[InlineKeyboardButton("🔙 Orqaga", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        f"```\n{daily_text}\n```",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def admin_manage_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin boshqaruvi callback"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.edit_message_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    # Adminlar ro'yxatini olish
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id, username FROM admins')
    admins = cursor.fetchall()
    conn.close()
    
    admin_text = "⚙️ **ADMINLAR RO'YXATI**\n\n"
    
    for i, admin in enumerate(admins, 1):
        admin_text += f"{i}. ID: {admin[0]} | @{admin[1] or "Noma'lum"}\n"
    
    admin_text += "\n📝 **Admin qo'shish:**\n"
    admin_text += "`/addadmin [user_id]` - Yangi admin qo'shish\n"
    admin_text += "`/removeadmin [user_id]` - Adminni olib tashlash\n"
    admin_text += "\n📊 **Statistika: **\n"
    admin_text += "`/stats` - To'liq statistika\n"  
    admin_text += "`/broadcast [xabar]` - Hamma foydalanuvchilarga xabar yuborish\n"
    
    # Orqaga qaytish tugmasi
    keyboard = [[InlineKeyboardButton("🔙 Orqaga", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        admin_text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def admin_export_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ma'lumotlarni yuklash callback"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.edit_message_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    export_text = "📤 **MA'LUMOTLARNI YUKLASH**\n\n"
    export_text += "Quyidagi buyruqlar orqali ma'lumotlarni yuklab olishingiz mumkin:\n\n"
    export_text += "📊 **Statistika fayllari:**\n"
    export_text += "`/export users` - Foydalanuvchilar ro'yxati (CSV)\n"
    export_text += "`/export stats` - Umumiy statistika (CSV)\n"
    export_text += "`/export daily` - Kunlik statistika (CSV)\n"
    
    # Orqaga qaytish tugmasi
    keyboard = [[InlineKeyboardButton("🔙 Orqaga", callback_data="admin_back")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        export_text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

async def admin_back_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin panelga qaytish"""
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    
    if not is_admin(user_id):
        await query.edit_message_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    # Statistikani yangilash
    update_daily_stats()
    
    # Bot statistikasini olish
    stats = get_bot_stats()
    
    # Kunlik statistikani olish
    daily_stats = get_daily_stats(7)
    
    # Admin panel keyboard
    keyboard = [
        [InlineKeyboardButton("📊 Umumiy statistika", callback_data="admin_stats")],
        [InlineKeyboardButton("👥 Foydalanuvchilar", callback_data="admin_users")],
        [InlineKeyboardButton("📈 Kunlik statistika", callback_data="admin_daily")],
        [InlineKeyboardButton("⚙️ Admin boshqaruvi", callback_data="admin_manage")],
        [InlineKeyboardButton("📤 Ma'lumotlarni yuklash", callback_data="admin_export")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    stats_text = f"""
🔧 **ADMIN PANELI**

👥 **Foydalanuvchilar:**
• Jami obunachilar: {stats['total_users']}
• Bugungi faol: {stats['today_active']}
• Bugungi yangi: {stats['today_new']}

📄 **PDF Statistika:**
• Jami PDF lar: {stats['total_pdfs']}
• Jami fayllar: {stats['total_files']}

📅 **Oxirgi 7 kun:**
"""
    
    # Oxirgi 7 kunlik statistikani qo'shish
    for day in daily_stats[:5]:
        day_date = datetime.strptime(day[0], '%Y-%m-%d').strftime('%d.%m')
        stats_text += f"• {day_date}: {day[2]} faol, {day[1]} yangi\n"
    
    await query.edit_message_text(
        stats_text,
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )

# ===== ADMIN COMMANDS =====
async def add_admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin qo'shish buyrug'i"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    if not context.args:
        await update.message.reply_text("❌ Foydalanish: `/addadmin [user_id]`")
        return
    
    try:
        new_admin_id = int(context.args[0])
        add_admin(new_admin_id, "Noma'lum")
        await update.message.reply_text(f"✅ {new_admin_id} admin sifatida qo'shildi!")
    except ValueError:
        await update.message.reply_text("❌ Noto'g'ri user_id format!")

async def remove_admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Adminni olib tashlash buyrug'i"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    if not context.args:
        await update.message.reply_text("❌ Foydalanish: `/removeadmin [user_id]`")
        return
    
    try:
        admin_id = int(context.args[0])
        remove_admin(admin_id)
        await update.message.reply_text(f"✅ {admin_id} adminlik huquqidan mahrum qilindi!")
    except ValueError:
        await update.message.reply_text("❌ Noto'g'ri user_id format!")

async def broadcast_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Barcha foydalanuvchilarga xabar yuborish"""
    user_id = update.effective_user.id
    
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    if not context.args:
        await update.message.reply_text("❌ Foydalanish: `/broadcast [xabar]`")
        return
    
    message = " ".join(context.args)
    
    # Foydalanuvchilar ro'yxatini olish
    conn = sqlite3.connect('data/bot_stats.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM users')
    users = cursor.fetchall()
    conn.close()
    
    sent = 0
    failed = 0
    
    # Har bir foydalanuvchiga xabar yuborish
    for user in users:
        try:
            await context.bot.send_message(
                chat_id=user[0],
                text=f"📢 **Bot yangiligi:**\n\n{message}"
            )
            sent += 1
        except:
            failed += 1
    
    await update.message.reply_text(
        f"📢 **Xabar yuborildi!**\n\n"
        f"✅ Muvaffaqiyatli: {sent}\n"
        f"❌ Muvaffaqiyatsiz: {failed}"
    )

# ===== PROGRESS XABARINI KO'RSATISH =====
async def show_progress_message(user_id: int, context: ContextTypes.DEFAULT_TYPE):
    """Progress xabarini ko'rsatish"""
    try:
        files = user_files.get(user_id, [])
        current_pdf_num = user_pdf_counter.get(user_id, 1)
        
        if not files:
            return
        
        total_files = len(files)
        
        # Progress bar yaratish
        progress_bar_length = 10
        filled = min(progress_bar_length, total_files)
        progress_bar = "🟩" * filled + "⬜" * (progress_bar_length - filled)
        progress_percentage = min(100, total_files * 10)
        
        progress_text = f"""
⏳ **Fayllar yuklanmoqda...**

{progress_bar}
📊 **Progress:** {progress_percentage}%
📁 **Fayllar soni:** {total_files} ta
🔢 **PDF:** #{current_pdf_num}

⏰ **PDF 3 soniyadan so'ng avtomatik yaratiladi...**
"""
        
        # Agar oldingi progress xabari bo'lsa, uni o'chirish
        if user_id in user_progress_msg_id:
            try:
                await context.bot.delete_message(
                    chat_id=user_id,
                    message_id=user_progress_msg_id[user_id]
                )
            except:
                pass
        
        # Yangi progress xabarini yuborish
        msg = await context.bot.send_message(
            chat_id=user_id,
            text=progress_text,
            parse_mode='Markdown'
        )
        user_progress_msg_id[user_id] = msg.message_id
        
    except Exception as e:
        print(f"Progress xabarini ko'rsatish xatosi: {e}")

# ===== PROGRESS XABARINI O'CHIRISH =====
async def delete_progress_message(user_id: int, context: ContextTypes.DEFAULT_TYPE):
    """Progress xabarini o'chirish"""
    try:
        if user_id in user_progress_msg_id:
            await context.bot.delete_message(
                chat_id=user_id,
                message_id=user_progress_msg_id[user_id]
            )
            user_progress_msg_id.pop(user_id, None)
    except:
        pass

# ===== FAYL QABUL QILISH =====
async def collect(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    user_id = u.id
    
    # Foydalanuvchi faolligini yangilash
    update_user_activity(user_id)
    
    # Obunani tekshirish
    if user_id not in user_subscribed or not user_subscribed[user_id]:
        is_subscribed = await check_subscription(user_id, context)
        if not is_subscribed:
            channel_info = await get_channel_info(context)
            keyboard = [
                [InlineKeyboardButton(f"📢 {channel_info['title']}", url=channel_info["link"])],
                [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_subscription")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"❌ Botdan foydalanish uchun kanalga obuna bo'lishingiz kerak!",
                reply_markup=reply_markup
            )
            return
        user_subscribed[user_id] = True
    
    # Agar user birinchi marta fayl yuborsa
    if user_id not in user_files:
        user_files[user_id] = []
        user_pdf_counter[user_id] = 1
    
    # Agar PDF allaqachon yaratilgan bo'lsa (fayllar bo'sh), yangi PDF raqamini oshirish
    if len(user_files.get(user_id, [])) == 0 and user_id in user_pdf_counter:
        user_pdf_counter[user_id] += 1
    
    files_added = []

    # Document
    if update.message.document:
        d = update.message.document
        f = await context.bot.get_file(d.file_id)
        safe_filename = d.file_name.replace("/", "_").replace("\\", "_")
        timestamp = datetime.now().strftime('%H%M%S_%f')[:-3]
        path = f"temp/{user_id}_{timestamp}_{safe_filename}"
        await f.download_to_drive(path)
        files_added.append(path)

    # Photo
    elif update.message.photo:
        p = update.message.photo[-1]
        f = await context.bot.get_file(p.file_id)
        timestamp = datetime.now().strftime('%H%M%S_%f')[:-3]
        path = f"temp/{user_id}_{timestamp}_{p.file_id}.jpg"
        await f.download_to_drive(path)

        img = Image.open(path).convert("RGB")
        img.thumbnail((2000, 2000))
        img.save(path, quality=80)
        files_added.append(path)

    if not files_added:
        await update.message.reply_text("❌ Fayl topilmadi!")
        return

    user_files[user_id].extend(files_added)
    
    # Progress xabarini ko'rsatish
    await show_progress_message(user_id, context)
    
    # Avvalgi taskni bekor qilish (agar bo'lsa)
    if user_id in user_tasks:
        try:
            user_tasks[user_id].cancel()
        except:
            pass

    # Yangi task yaratish (3 soniya)
    user_tasks[user_id] = asyncio.create_task(delayed_pdf(user_id, context))

# ===== 3 SONIYA KUTIB PDF YARATISH =====
async def delayed_pdf(user_id, context):
    try:
        # 3 soniya kutish
        await asyncio.sleep(3)

        files = user_files.get(user_id)
        if not files or len(files) == 0:
            user_tasks.pop(user_id, None)
            await delete_progress_message(user_id, context)
            return

        # Progress xabarini o'chirish
        await delete_progress_message(user_id, context)

        # PDF yaratish xabarini yuborish
        creating_msg = await context.bot.send_message(
            chat_id=user_id,
            text="🔄 **PDF yaratilmoqda...**\n\nIltimos, kuting...",
            parse_mode='Markdown'
        )

        await create_and_send_pdf(user_id, context)
        
        # PDF yaratish xabarini o'chirish
        try:
            await context.bot.delete_message(
                chat_id=user_id,
                message_id=creating_msg.message_id
            )
        except:
            pass
        
    except asyncio.CancelledError:
        await delete_progress_message(user_id, context)
        return
    except Exception as e:
        print(f"delayed_pdf xatosi: {e}")
        await delete_progress_message(user_id, context)
    finally:
        user_tasks.pop(user_id, None)

# ===== UNICODE MATNLAR UCHUN FONTLAR =====
def add_unicode_support_to_pdf(pdf):
    """PDF ga Unicode support qo'shish"""
    # DejaVu Unicode fontlarini qo'shish (agar mavjud bo'lsa)
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "./fonts/DejaVuSans.ttf"
    ]
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdf.add_font('DejaVu', '', font_path, uni=True)
                pdf.set_font('DejaVu', '', 12)
                return True
            except:
                continue
    
    # Agar Unicode font topilmasa, Arial ni ishlatish
    try:
        pdf.add_font('Arial', '', 'arial.ttf', uni=True)
        pdf.set_font('Arial', '', 12)
        return True
    except:
        # Eng oxirgi chora - standart font
        pdf.set_font("Arial", size=12)
        return False

# ===== MATNNI TO'G'RI FORMATLASH =====
def clean_text(text):
    """Matnni tozalash va encoding muammolarini hal qilish"""
    if not text:
        return ""
    
    # Encoding muammolarini hal qilish
    try:
        # UTF-8 ga decode qilish
        if isinstance(text, bytes):
            text = text.decode('utf-8')
        
        # Maxsus belgilarni almashtirish
        replacements = {
            'â€™': "'",
            'â€"': "-",
            'â€"': '"',
            'â€˜': "'",
            'â€™': "'",
            'â€"': '"',
            'â€"': '"',
            'â€"': '-',
            'â€"': '--',
            '\x00': '',  # Null belgilar
            '\r\n': '\n',  # Windows newline
            '\r': '\n',    # Mac newline
        }
        
        for old, new in replacements.items():
            text = text.replace(old, new)
        
        # Bo'sh joylarni tozalash
        text = text.strip()
        
        # Agar matn bo'sh bo'lsa
        if not text or text.isspace():
            return ""
            
        return text
        
    except Exception as e:
        print(f"Matnni tozalash xatosi: {e}")
        # Agar encoding muammosi bo'lsa, oddiy matn qaytarish
        try:
            return str(text).encode('ascii', 'ignore').decode('ascii')
        except:
            return ""

# ===== DOCX FAYLLARINI QAYTA ISHLASH =====
def process_docx_file(path):
    """DOCX faylini o'qish va to'g'ri formatlash"""
    try:
        doc = docx.Document(path)
        paragraphs = []
        
        for para in doc.paragraphs:
            text = para.text
            cleaned_text = clean_text(text)
            if cleaned_text:  # Faqat bo'sh bo'lmagan paragraphlar
                paragraphs.append(cleaned_text)
        
        return paragraphs
        
    except Exception as e:
        print(f"DOCX faylini o'qish xatosi: {e}")
        return []

# ===== EXCEL FAYLLARINI QAYTA ISHLASH =====
def process_excel_file(path):
    """Excel faylini o'qish va to'g'ri formatlash"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        data = []
        
        for row in ws.iter_rows(values_only=True):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    # Cell ma'lumotini to'g'ri formatlash
                    cell_text = str(cell)
                    cleaned_text = clean_text(cell_text)
                    row_data.append(cleaned_text)
            data.append(row_data)
        
        return data
        
    except Exception as e:
        print(f"Excel faylini o'qish xatosi: {e}")
        return []

# ===== PDF YARATISH VA YUBORISH =====
async def create_and_send_pdf(user_id, context):
    try:
        files = user_files.get(user_id)
        if not files or len(files) == 0:
            return

        pdf = FPDF()
        
        # Unicode support qo'shish
        unicode_supported = add_unicode_support_to_pdf(pdf)
        if not unicode_supported:
            pdf.set_font("Arial", size=12)
        
        pdf.set_auto_page_break(True, 10)
        
        current_pdf_num = user_pdf_counter.get(user_id, 1)
        total_files = len(files)

        for idx, path in enumerate(files, 1):
            try:
                ext = os.path.splitext(path)[1].lower()
                
                if ext in [".jpg", ".jpeg", ".png"]:
                    # RASM uchun
                    pdf.add_page()
                    try:
                        img = Image.open(path)
                        w, h = img.size
                        
                        # A4 formatiga moslashtirish
                        page_width = 190  # mm
                        page_height = 277  # mm
                        
                        ratio = min(page_width / w, page_height / h)
                        new_width = w * ratio
                        new_height = h * ratio
                        
                        # Markazga joylashtirish
                        x = (210 - new_width) / 2
                        y = (297 - new_height) / 2
                        
                        pdf.image(path, x=x, y=y, w=new_width, h=new_height)
                        
                    except Exception as img_e:
                        print(f"Rasm xatosi: {img_e}")
                        pdf.add_page()
                        pdf.cell(0, 10, f"Rasmni ochishda xatolik: {path}", 0, 1)

                elif ext == ".docx":
                    # WORD DOCX uchun
                    try:
                        paragraphs = process_docx_file(path)
                        
                        if paragraphs:
                            pdf.add_page()
                            pdf.set_font_size(12)
                            
                            for para in paragraphs:
                                # Matnni PDF ga qo'shish
                                try:
                                    # Ko'p qatorli matn
                                    pdf.multi_cell(0, 8, para)
                                    pdf.ln(4)
                                except Exception as write_e:
                                    print(f"Matn yozish xatosi: {write_e}")
                                    # Agar xatolik bo'lsa, encoding ni o'zgartirish
                                    try:
                                        safe_text = para.encode('latin-1', 'replace').decode('latin-1')
                                        pdf.multi_cell(0, 8, safe_text)
                                        pdf.ln(4)
                                    except:
                                        pdf.multi_cell(0, 8, "[Matnni ko'rsatish mumkin emas]")
                                        pdf.ln(4)
                        else:
                            pdf.add_page()
                            pdf.cell(0, 10, f"DOCX fayl bo'sh yoki o'qish mumkin emas: {path}", 0, 1)
                            
                    except Exception as doc_e:
                        print(f"Word xatosi: {doc_e}")
                        pdf.add_page()
                        pdf.cell(0, 10, f"DOCX faylni qayta ishlashda xatolik", 0, 1)

                elif ext in [".xlsx", ".xls"]:
                    # EXCEL uchun
                    try:
                        data = process_excel_file(path)
                        
                        if data:
                            pdf.add_page()
                            pdf.set_font_size(10)
                            
                            for row_idx, row in enumerate(data):
                                # Har bir satrni bitta qatorda chiqarish
                                row_text = " | ".join([str(cell) for cell in row])
                                
                                # Juda uzun satrlarni qisqartirish
                                if len(row_text) > 150:
                                    row_text = row_text[:147] + "..."
                                
                                try:
                                    pdf.multi_cell(0, 6, row_text)
                                except:
                                    # Encoding muammosi bo'lsa
                                    try:
                                        safe_text = row_text.encode('latin-1', 'replace').decode('latin-1')
                                        pdf.multi_cell(0, 6, safe_text)
                                    except:
                                        pdf.multi_cell(0, 6, f"[Satr {row_idx + 1}]")
                        else:
                            pdf.add_page()
                            pdf.cell(0, 10, f"Excel fayl bo'sh yoki o'qish mumkin emas", 0, 1)
                            
                    except Exception as excel_e:
                        print(f"Excel xatosi: {excel_e}")
                        pdf.add_page()
                        pdf.cell(0, 10, f"Excel faylni qayta ishlashda xatolik", 0, 1)
                        
                else:
                    # Boshqa fayl turlari uchun
                    pdf.add_page()
                    pdf.cell(0, 10, f"Noma'lum fayl turi: {ext}", 0, 1)
                    
            except Exception as e:
                print(f"Fayl qayta ishlash xatosi: {e}")
                pdf.add_page()
                pdf.cell(0, 10, f"Faylni qayta ishlashda xatolik: {path}", 0, 1)
                continue

        # PDF ni saqlash
        out = f"temp/{user_id}_pdf_{current_pdf_num}.pdf"
        pdf.output(out)

        # Foydalanuvchiga yuborish
        try:
            with open(out, "rb") as pdf_file:
                await context.bot.send_document(
                    chat_id=user_id,
                    document=pdf_file,
                    filename=f"PDF_{current_pdf_num}.pdf",
                    caption=f"✅ **PDF #{current_pdf_num} tayyor!**\n\n"
                           f"📊 {total_files} ta fayl birlashtirildi.\n\n"
                           f"Yangi fayllar yuborishingiz mumkin!",
                    parse_mode='Markdown'
                )
        except Exception as send_e:
            print(f"Yuborish xatosi: {send_e}")
            await context.bot.send_message(
                chat_id=user_id,
                text=f"❌ PDF yuborishda xatolik"
            )

        # Fayllarni tozalash
        for f in files:
            try:
                if os.path.exists(f):
                    os.remove(f)
            except:
                pass
        
        # Chiqish faylini o'chirish
        try:
            if os.path.exists(out):
                os.remove(out)
        except:
            pass
        
        # Fayllar ro'yxatini tozalash
        user_files[user_id] = []
        
        # PDF raqamini oshirish
        user_pdf_counter[user_id] = current_pdf_num + 1
        
        # Statistikani yangilash
        increment_user_stats(user_id, pdfs=1, files=total_files)
        
    except Exception as e:
        print(f"PDF yaratish xatosi: {e}")
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"❌ PDF yaratishda xatolik: {str(e)[:100]}"
            )
        except:
            pass

# ===== /newpdf - DARHOL PDF YARATISH =====
async def new_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    user_id = u.id
    
    # Foydalanuvchi faolligini yangilash
    update_user_activity(user_id)
    
    # Obunani tekshirish
    if user_id not in user_subscribed or not user_subscribed[user_id]:
        is_subscribed = await check_subscription(user_id, context)
        if not is_subscribed:
            channel_info = await get_channel_info(context)
            keyboard = [
                [InlineKeyboardButton(f"📢 {channel_info['title']}", url=channel_info["link"])],
                [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_subscription")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"❌ Botdan foydalanish uchun kanalga obuna bo'lishingiz kerak!",
                reply_markup=reply_markup
            )
            return
        user_subscribed[user_id] = True
    
    # Agar fayllar bo'lsa, PDF yaratish
    if user_files.get(user_id) and len(user_files[user_id]) > 0:
        # Avvalgi taskni bekor qilish
        if user_id in user_tasks:
            try:
                user_tasks[user_id].cancel()
            except:
                pass
            user_tasks.pop(user_id, None)
        
        # Progress xabarini o'chirish
        await delete_progress_message(user_id, context)
        
        # Darhol PDF yaratish
        await create_and_send_pdf(user_id, context)
    else:
        await update.message.reply_text(
            "📄 **PDF yaratish uchun avval fayllar yuboring!**",
            parse_mode='Markdown'
        )

# ===== /clean - TOZALASH =====
async def clean(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    user_id = u.id
    
    # Foydalanuvchi faolligini yangilash
    update_user_activity(user_id)
    
    # Obunani tekshirish
    if user_id not in user_subscribed or not user_subscribed[user_id]:
        is_subscribed = await check_subscription(user_id, context)
        if not is_subscribed:
            channel_info = await get_channel_info(context)
            keyboard = [
                [InlineKeyboardButton(f"📢 {channel_info['title']}", url=channel_info["link"])],
                [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_subscription")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"❌ Botdan foydalanish uchun kanalga obuna bo'lishingiz kerak!",
                reply_markup=reply_markup
            )
            return
        user_subscribed[user_id] = True
    
    # Taskni bekor qilish
    if user_id in user_tasks:
        try:
            user_tasks[user_id].cancel()
        except:
            pass
        user_tasks.pop(user_id, None)
    
    # Progress xabarini o'chirish
    await delete_progress_message(user_id, context)
    
    # Fayllarni tozalash
    if user_id in user_files:
        for f in user_files[user_id]:
            try:
                if os.path.exists(f):
                    os.remove(f)
            except:
                pass
        user_files[user_id] = []
    
    # PDF counter ni reset qilish
    user_pdf_counter[user_id] = 1
    
    await update.message.reply_text(
        "✅ **Barcha fayllar tozalandi!**",
        parse_mode='Markdown'
    )

# ===== OBUNANI TEKSHIRISH CALLBACK =====
async def check_subscription_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    user_id = query.from_user.id
    channel_info = await get_channel_info(context)
    
    is_subscribed = await check_subscription(user_id, context)
    
    if is_subscribed:
        user_subscribed[user_id] = True
        
        # Reset everything
        user_files[user_id] = []
        user_pdf_counter[user_id] = 1
        user_progress_msg_id.pop(user_id, None)
        
        await query.edit_message_text(
            f"✅ **Obuna tasdiqlandi!**\n\n"
            f"Endi fayllarni yuborishingiz mumkin.",
            parse_mode='Markdown'
        )
    else:
        keyboard = [
            [InlineKeyboardButton(f"📢 {channel_info['title']}", url=channel_info["link"])],
            [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_subscription")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            f"❌ **Hali obuna bo'lmagansiz!**\n\n"
            f"Kanalga obuna bo'ling:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )

# ===== /status - HOLATNI TEKSHIRISH =====
async def status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    u = update.effective_user
    user_id = u.id
    
    # Foydalanuvchi faolligini yangilash
    update_user_activity(user_id)
    
    # Obunani tekshirish
    if user_id not in user_subscribed or not user_subscribed[user_id]:
        is_subscribed = await check_subscription(user_id, context)
        if not is_subscribed:
            channel_info = await get_channel_info(context)
            keyboard = [
                [InlineKeyboardButton(f"📢 {channel_info['title']}", url=channel_info["link"])],
                [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="check_subscription")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"❌ Botdan foydalanish uchun kanalga obuna bo'lishingiz kerak!",
                reply_markup=reply_markup
            )
            return
        user_subscribed[user_id] = True
    
    if user_id in user_files:
        files_count = len(user_files[user_id])
        current_pdf = user_pdf_counter.get(user_id, 1)
        
        if files_count > 0:
            await update.message.reply_text(
                f"📊 **Holat:**\n"
                f"• Keyingi PDF raqami: **#{current_pdf}**\n"
                f"• Fayllar soni: **{files_count}**\n"
                f"• PDF 3 soniyadan so'ng yaratiladi\n\n"
                f"📎 **/newpdf** - Darhol PDF yaratish",
                parse_mode='Markdown'
            )
        else:
            await update.message.reply_text(
                f"📊 **Holat:**\n"
                f"• Keyingi PDF raqami: **#{current_pdf}**\n"
                f"• Fayllar soni: **0**\n\n"
                f"✅ Fayl yuborish uchun tayyor!",
                parse_mode='Markdown'
            )
    else:
        await update.message.reply_text(
            "📄 **Fayl yuborish uchun tayyor!**",
            parse_mode='Markdown'
        )

# ===== MAIN =====
def main():
    app = ApplicationBuilder().token(TOKEN).build()
    
    # Command handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("newpdf", new_pdf))
    app.add_handler(CommandHandler("clean", clean))
    app.add_handler(CommandHandler("status", status))
    app.add_handler(CommandHandler("admin", admin_panel))
    app.add_handler(CommandHandler("addadmin", add_admin_command))
    app.add_handler(CommandHandler("removeadmin", remove_admin_command))
    app.add_handler(CommandHandler("broadcast", broadcast_command))
    
    # Callback handlers
    app.add_handler(CallbackQueryHandler(check_subscription_callback, pattern="check_subscription"))
    app.add_handler(CallbackQueryHandler(admin_stats_callback, pattern="admin_stats"))
    app.add_handler(CallbackQueryHandler(admin_users_callback, pattern="admin_users"))
    app.add_handler(CallbackQueryHandler(admin_daily_callback, pattern="admin_daily"))
    app.add_handler(CallbackQueryHandler(admin_manage_callback, pattern="admin_manage"))
    app.add_handler(CallbackQueryHandler(admin_export_callback, pattern="admin_export"))
    app.add_handler(CallbackQueryHandler(admin_back_callback, pattern="admin_back"))
    
    # Message handlers
    app.add_handler(MessageHandler(filters.Document.ALL | filters.PHOTO, collect))
    
    print("🤖 Bot ishga tushdi")

    app = Flask(__name__)

    application = ApplicationBuilder().token(TOKEN).build()

    # ===== HANDLERLARNI QO‘SHISH =====
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("admin", admin_panel))
    application.add_handler(CommandHandler("newpdf", new_pdf))
    application.add_handler(CommandHandler("clean", clean))
    application.add_handler(CommandHandler("addadmin", add_admin_command))
    application.add_handler(CommandHandler("removeadmin", remove_admin_command))
    application.add_handler(CommandHandler("broadcast", broadcast_command))

    application.add_handler(
        MessageHandler(
            filters.Document.ALL | filters.PHOTO,
            collect
        )
    )

    application.add_handler(CallbackQueryHandler(admin_stats_callback, pattern="admin_stats"))
    application.add_handler(CallbackQueryHandler(admin_users_callback, pattern="admin_users"))
    application.add_handler(CallbackQueryHandler(admin_daily_callback, pattern="admin_daily"))
    application.add_handler(CallbackQueryHandler(admin_manage_callback, pattern="admin_manage"))
    application.add_handler(CallbackQueryHandler(admin_export_callback, pattern="admin_export"))
    application.add_handler(CallbackQueryHandler(admin_back_callback, pattern="admin_back"))
    application.add_handler(CallbackQueryHandler(lambda u, c: start(u, c), pattern="check_subscription"))


    @app.route("/", methods=["GET"])
    def home():
        return "Telegram bot Web Service ishlayapti ✅"


    @app.route("/webhook", methods=["POST"])
    async def webhook():
        update = Update.de_json(request.get_json(force=True), application.bot)
        await application.process_update(update)
        return "ok"


    if __name__ == "__main__":
        WEBHOOK_URL = os.environ.get("RENDER_EXTERNAL_URL")
        application.bot.set_webhook(f"{WEBHOOK_URL}/webhook")
        app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))


if __name__ == "__main__":
    main()